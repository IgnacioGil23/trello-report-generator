import os
import sys
import requests
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime, timedelta, timezone

# --- Lógica de Trello ---

def load_env_vars():
    """Carga y valida las variables de entorno de Trello."""
    load_dotenv()
    api_key = os.getenv("TRELLO_API_KEY")
    token = os.getenv("TRELLO_TOKEN")
    board_id = os.getenv("TRELLO_BOARD_ID")
    theme = os.getenv("APP_THEME", "blue")
    
    creds = (api_key, token, board_id)
    
    if not all(creds):
        return None, theme
        
    return creds, theme

def get_lists(api_key, token, board_id):
    """Obtiene todas las listas (etapas) del tablero."""
    try:
        url = f"https://api.trello.com/1/boards/{board_id}/lists"
        params = {"key": api_key, "token": token}
        response = requests.get(url, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        raise ConnectionError(f"Error al obtener listas: {e}")

def get_cards(api_key, token, board_id):
    """Obtiene todas las tarjetas (clientes) del tablero."""
    try:
        url = f"https://api.trello.com/1/boards/{board_id}/cards"
        params = {"key": api_key, "token": token}
        response = requests.get(url, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        raise ConnectionError(f"Error al obtener tarjetas: {e}")

def get_card_actions(api_key, token, card_id):
    """Obtiene el historial de acciones de una tarjeta."""
    try:
        url = f"https://api.trello.com/1/cards/{card_id}/actions"
        params = {
            "key": api_key, "token": token,
            "filter": "updateCard:idList,createCard", "limit": 1000
        }
        response = requests.get(url, params=params)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException:
        return []

def format_trello_date(date_str):
    """Convierte fecha de Trello a formato local y legible."""
    if not date_str:
        return None
    try:
        utc_time = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        local_time = utc_time.astimezone(timezone.utc).astimezone()
        return local_time.strftime('%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        return date_str

def parse_actions(actions, list_id_to_name):
    """Analiza las acciones para determinar el tiempo en cada etapa."""
    if not actions:
        return []
    
    events = []
    for action in actions:
        date = action['date']
        if action['type'] == 'createCard':
            list_id = action['data']['list']['id']
            events.append({'date': date, 'list_id': list_id, 'type': 'creation'})
        elif action['type'] == 'updateCard' and 'listAfter' in action['data']:
            list_id = action['data']['listAfter']['id']
            events.append({'date': date, 'list_id': list_id, 'type': 'move'})

    if not events:
        return []

    events_sorted = sorted(events, key=lambda x: x['date'])
    
    etapas = []
    for i, event in enumerate(events_sorted):
        fecha_entrada = event['date']
        fecha_salida = events_sorted[i + 1]['date'] if i + 1 < len(events_sorted) else None
        
        etapas.append({
            'list_id': event['list_id'],
            'etapa': list_id_to_name.get(event['list_id'], 'Desconocida'),
            'fecha_entrada': fecha_entrada,
            'fecha_salida': fecha_salida
        })
    return etapas

def filter_by_date_range(etapas, start_date, end_date):
    """Filtra etapas por rango de fechas."""
    if start_date is None and end_date is None:
        return etapas
    
    filtered_etapas = []
    for etapa in etapas:
        try:
            fecha_entrada = datetime.fromisoformat(etapa['fecha_entrada'].replace('Z', '+00:00')).astimezone(timezone.utc).astimezone()
            if start_date and fecha_entrada < start_date:
                continue
            if end_date and fecha_entrada > end_date:
                continue
            filtered_etapas.append(etapa)
        except:
            continue
    return filtered_etapas

# --- Generadores de Reportes ---

def generate_detailed_report(cards, list_id_to_name, api_key, token, start_date=None, end_date=None, progress_callback=None):
    report_rows = []
    total_cards = len(cards)
    
    for i, card in enumerate(cards):
        actions = get_card_actions(api_key, token, card['id'])
        etapas = parse_actions(actions, list_id_to_name)
        
        if start_date or end_date:
            etapas = filter_by_date_range(etapas, start_date, end_date)

        if not etapas and card.get('idList'):
            etapas = [{'list_id': card['idList'], 'etapa': list_id_to_name.get(card['idList'], 'Desconocida'),
                       'fecha_entrada': card.get('dateLastActivity'), 'fecha_salida': None}]

        for etapa in etapas:
            tiempo_dias = None
            if etapa['fecha_salida'] and etapa['fecha_entrada']:
                try:
                    t_in = datetime.fromisoformat(etapa['fecha_entrada'].replace('Z', '+00:00'))
                    t_out = datetime.fromisoformat(etapa['fecha_salida'].replace('Z', '+00:00'))
                    tiempo_dias = round((t_out - t_in).total_seconds() / 86400, 2)
                except ValueError:
                    tiempo_dias = None
            
            report_rows.append({
                'Cliente': card['name'], 'Etapa': etapa['etapa'],
                'Fecha de entrada': format_trello_date(etapa['fecha_entrada']),
                'Fecha de salida': format_trello_date(etapa['fecha_salida']),
                'Tiempo en etapa (días)': tiempo_dias
            })
        
        if progress_callback:
            progress_callback((i + 1) / total_cards)
            
    return pd.DataFrame(report_rows)

def generate_time_analysis_report(cards, list_id_to_name, api_key, token, progress_callback=None):
    all_times = []
    total_cards = len(cards)
    for i, card in enumerate(cards):
        actions = get_card_actions(api_key, token, card['id'])
        etapas = parse_actions(actions, list_id_to_name)
        for etapa in etapas:
            if etapa['fecha_salida'] and etapa['fecha_entrada']:
                try:
                    t_in = datetime.fromisoformat(etapa['fecha_entrada'].replace('Z', '+00:00'))
                    t_out = datetime.fromisoformat(etapa['fecha_salida'].replace('Z', '+00:00'))
                    all_times.append({'Etapa': etapa['etapa'], 'Tiempo (días)': (t_out - t_in).total_seconds() / 86400})
                except:
                    continue
        if progress_callback:
            progress_callback((i + 1) / total_cards)
            
    if not all_times:
        return pd.DataFrame()
    
    df = pd.DataFrame(all_times)
    return df.groupby('Etapa')['Tiempo (días)'].agg(['mean', 'min', 'max', 'count']).round(2).reset_index()

def generate_movement_report(cards, list_id_to_name, api_key, token, start_date=None, end_date=None, progress_callback=None):
    movements = []
    total_cards = len(cards)
    for i, card in enumerate(cards):
        actions = get_card_actions(api_key, token, card['id'])
        etapas = parse_actions(actions, list_id_to_name)
        
        if start_date or end_date:
            etapas = filter_by_date_range(etapas, start_date, end_date)
            
        for j in range(len(etapas) - 1):
            movements.append({'De': etapas[j]['etapa'], 'A': etapas[j+1]['etapa'], 'Cliente': card['name']})
        
        if progress_callback:
            progress_callback((i + 1) / total_cards)
            
    if not movements:
        return pd.DataFrame()
    
    df = pd.DataFrame(movements)
    return df.groupby(['De', 'A']).size().reset_index(name='Cantidad')

def generate_current_status_report(cards, list_id_to_name):
    current_status = [{'Cliente': card['name'], 'Etapa Actual': list_id_to_name.get(card.get('idList'), 'Desconocida'),
                       'Última Actividad': format_trello_date(card.get('dateLastActivity'))} for card in cards]
    df = pd.DataFrame(current_status)
    summary = df.groupby('Etapa Actual').size().reset_index(name='Cantidad')
    return df, summary

def generate_velocity_report(cards, list_id_to_name, api_key, token, progress_callback=None):
    client_times = []
    total_cards = len(cards)
    for i, card in enumerate(cards):
        actions = get_card_actions(api_key, token, card['id'])
        etapas = parse_actions(actions, list_id_to_name)
        if len(etapas) > 1:
            try:
                t_in = datetime.fromisoformat(etapas[0]['fecha_entrada'].replace('Z', '+00:00'))
                t_out = datetime.fromisoformat(etapas[-1]['fecha_salida'].replace('Z', '+00:00')) if etapas[-1]['fecha_salida'] else datetime.now(timezone.utc)
                client_times.append({'Cliente': card['name'], 'Tiempo Total (días)': round((t_out - t_in).total_seconds() / 86400, 2),
                                     'Etapas Completadas': len([e for e in etapas if e['fecha_salida']])})
            except:
                continue
        if progress_callback:
            progress_callback((i + 1) / total_cards)
            
    if not client_times:
        return pd.DataFrame()
    
    return pd.DataFrame(client_times).sort_values('Tiempo Total (días)')

# --- Formato y Guardado Excel ---

def adjust_column_widths(worksheet, dataframe):
    """Ajusta automáticamente el ancho de las columnas."""
    for idx, col in enumerate(dataframe.columns):
        max_length = max(len(str(col)), dataframe[col].astype(str).map(len).max())
        worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)

def apply_excel_formatting(worksheet, dataframe):
    """Aplica formato profesional al archivo Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin', color='D1D3D4'), right=Side(style='thin', color='D1D3D4'),
                    top=Side(style='thin', color='D1D3D4'), bottom=Side(style='thin', color='D1D3D4'))

    for col_idx, column in enumerate(dataframe.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center'); cell.border = border
    
    for row_idx in range(2, len(dataframe) + 2):
        for col_idx in range(1, len(dataframe.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = cell_font; cell.border = border

def create_time_analysis_chart(worksheet, df):
    """Crea un gráfico de barras para el análisis de tiempos."""
    from openpyxl.chart import BarChart, Reference
    
    chart = BarChart()
    chart.title = "Tiempo Promedio por Etapa"
    chart.y_axis.title = "Etapa"
    chart.x_axis.title = "Días Promedio"
    
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(df) + 1, max_col=2)
    cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(df) + 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    
    worksheet.add_chart(chart, "G2")

def create_status_summary_chart(worksheet, df):
    """Crea un gráfico de torta para el resumen de estado actual."""
    from openpyxl.chart import PieChart, Reference
    
    chart = PieChart()
    chart.title = "Distribución de Tarjetas por Etapa"
    
    labels = Reference(worksheet, min_col=1, min_row=2, max_row=len(df) + 1)
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(df) + 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    
    worksheet.add_chart(chart, "E2")

def save_report_to_excel(reports_dict, filename):
    """Guarda los reportes en un archivo Excel con formato y gráficos."""
    from openpyxl import load_workbook
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in reports_dict.items():
            if isinstance(df, tuple):
                for i, sub_df in enumerate(df):
                    sub_sheet_name = f"{sheet_name}_{i+1}" if i > 0 else sheet_name
                    sub_df.to_excel(writer, sheet_name=sub_sheet_name[:31], index=False)
            else:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    
    workbook = load_workbook(filename)
    for sheet_name, df in reports_dict.items():
        if isinstance(df, tuple):
            for i, sub_df in enumerate(df):
                sub_sheet_name = f"{sheet_name}_{i+1}" if i > 0 else sheet_name
                if sub_sheet_name[:31] in workbook.sheetnames:
                    worksheet = workbook[sub_sheet_name[:31]]
                    adjust_column_widths(worksheet, sub_df)
                    apply_excel_formatting(worksheet, sub_df)
        else:
            if sheet_name[:31] in workbook.sheetnames:
                worksheet = workbook[sheet_name[:31]]
                adjust_column_widths(worksheet, df)
                apply_excel_formatting(worksheet, df)
                
                # Añadir gráficos específicos
                if sheet_name == "Tiempos":
                    create_time_analysis_chart(worksheet, df)
                elif sheet_name == "Estado_Resumen":
                    create_status_summary_chart(worksheet, df)

    workbook.save(filename)
    workbook.close()

# --- Funciones para la ventana de configuración ---

def save_env_vars(api_key, token, board_id, theme="blue"):
    """Guarda las credenciales y el tema en el archivo .env."""
    with open(".env", "w") as f:
        f.write(f"TRELLO_API_KEY={api_key}\n")
        f.write(f"TRELLO_TOKEN={token}\n")
        f.write(f"TRELLO_BOARD_ID={board_id}\n")
        f.write(f"APP_THEME={theme}\n")

def test_trello_connection(api_key, token, board_id):
    """Prueba la conexión con Trello usando las credenciales proporcionadas."""
    try:
        url = f"https://api.trello.com/1/boards/{board_id}"
        params = {"key": api_key, "token": token, "fields": "name"}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        board_name = response.json().get("name")
        return True, f"Conectado al tablero '{board_name}'"
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            return False, "API Key o Token inválidos."
        elif e.response.status_code == 404:
            return False, "Board ID no encontrado."
        return False, f"Error HTTP: {e.response.status_code}"
    except requests.exceptions.RequestException as e:
        return False, f"Error de conexión: {e}"
    except Exception as e:
        return False, f"Error inesperado: {e}"
